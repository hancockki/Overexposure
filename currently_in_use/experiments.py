import driver
import view
import sys
import graph_creation
import create_graph_from_file as cff
import networkx as nx
import openpyxl
import re
FL_PREFIX = "test_files/"
FILE_DIRECTORY_PREFIX = "currently_in_use/"
RECORD_FILENAME = "all_cluster_data.txt"

def is_ID_reset():
    filename = FILE_DIRECTORY_PREFIX + "ID_tracker.txt"
    file = open(filename, 'r')
    # get current ID that is recorded in file
    if file.mode == 'r':
        ID = file.read()

    return ID == "1"

def generate_and_save_graphs(node_sizes, runs, m=2,p=0.2):
    if not is_ID_reset():
        do_stop = input("The ID counter is not reset, is this intentional? Data may be overwritten if you continue. Enter [Y] or [N]: ")
        if do_stop == "Y":
            sys.exit()
    for num_nodes in node_sizes:
        for i in range(runs):
            original_graphs, original_types = driver.generate_original_graphs(num_nodes, m, p)
            for O, graph_type in zip(original_graphs, original_types):
                ID = view.generate_ID()
                location = view.save_original(O, "n/a", "n/a", graph_type, ID, "n/a", "n/a")

def test_dif_combos(node_sizes, node_offset, runs, k_vals, appeals):
    graph_types = ["BA","ER","WS"]
    offset = [1,2,3]

    for graph, ID_off in zip(graph_types, offset):
        for appeal in appeals:
            for num_nodes_offset, num_nodes in zip(node_offset, node_sizes):
                for k in k_vals:
                    for i in range(runs):
                        filename = graph + "/" + str(num_nodes) + "/" + str((num_nodes_offset * runs * 3) + (i * 3) + ID_off)
                        driver.test_file(filename, k, appeal)

def record_cluster_data(node_sizes, node_offset, runs, k_vals, appeals):
    graph_types = ["BA","ER","WS"]
    offset = [1,2,3]
    for graph, ID_off in zip(graph_types, offset):
        for appeal in appeals:
            for num_nodes_offset, num_nodes in zip(node_offset, node_sizes):
                for k in k_vals:
                    for i in range(runs):
                        filename = graph + "/" + str(num_nodes) + "/" + str((num_nodes_offset * runs * 3) + (i * 3) + ID_off)
                        with open(FILE_DIRECTORY_PREFIX + RECORD_FILENAME, 'a') as graph_info:
                            graph_info.write('#' + filename + "/" + str(appeal) + '\n')
                            graph_info.close()
                        driver.test_file(filename, k, appeal, 'true')

def get_all_possible_cluster_sizes(node_sizes, runs, appeals):
    graph_types = ["BA","ER","WS"]
    offset = [1,2,3]
    
    possible_cluster_sizes = set()
    for graph, ID_off in zip(graph_types, offset):
        for appeal in appeals:
            appeal = float(appeal)
            for num_nodes_offset, num_nodes in enumerate(node_sizes):
                num_nodes_offset = 3
                for i in range(runs):
                    filename = "test_files/" + graph + "/" + str(num_nodes) + "/" + str((num_nodes_offset * runs * 3) + (i * 3) + ID_off) + ".txt"
                    print("Opening " + filename)
                    file_k, file_appeal, graph_type, ID, file_remove_cycles, file_assumption_1, O = cff.create_from_file(filename)
                    C = graph_creation.create_cluster_graph(O, appeal)
                    B = graph_creation.create_bipartite_from_cluster(C)
                    for node in B.nodes():
                        if B.nodes[node]['bipartite'] == 0:
                            possible_cluster_sizes.add(B.nodes[node]['weight'])
                    print("Finished recording clusters for " + filename)
                    B.clear()
                    C.clear()
                    O.clear()
    possible_cluster_sizes = list(possible_cluster_sizes)
    possible_cluster_sizes.sort(reverse=True)
    print("======================================")
    print("There are " + str(len(possible_cluster_sizes)) + " unique cluster sizes")

    # record all cluster sizes into a file
    filename = FILE_DIRECTORY_PREFIX + "all_cluster_sizes.txt"
    # put number on new line everytime
    with open(filename, 'w') as write_to_file:
        for size in possible_cluster_sizes:
            write_to_file.write(str(size) + "\n")
    print("Finished recording in file")

def record_occurences_of_cluster_sizes():
    all_cluster_data_filename = FILE_DIRECTORY_PREFIX + RECORD_FILENAME
    all_keys = get_all_possible_keys(all_cluster_data_filename)

    file = open(all_cluster_data_filename,"r")
    if file.mode == 'r':
        contents = file.read()
    lines = re.split("\n", contents)
    file.close()

    write_header(all_keys)
    
    for i, line in enumerate(lines):
        if line != '' and line[0] == '#':
            if i != 0:
                print("Finished Count for " + filename)
                save_cluster_size_data(num_nodes, graph, appeal, filename, cluster_sizes)
            cluster_sizes = reset_dict_from_list_keys(all_keys)
            graph_info = line.split('/')
            graph = graph_info[0][1:]
            num_nodes = graph_info[1]
            appeal = graph_info[3]
            filename = "test_files/" + graph + "/" + str(num_nodes) + "/" + str(graph_info[2]) + ".txt"
            print("Recording " + filename)
        else:
            cluster_sizes[int(line)] = cluster_sizes[int(line)] + 1

def get_all_possible_keys(filename):
    file = open(filename,"r")
    if file.mode == 'r':
        contents = file.read()
    lines = re.split("\n", contents)
    all_cluster_sizes = set()
    for line in lines:
        if line != '' and line[0] != '#':
            all_cluster_sizes.add(int(line))
    file.close()
    sorted = list(all_cluster_sizes)
    sorted.sort()
    return sorted

def reset_dict_from_list_keys(keys):
    cluster_sizes = dict()
    for key in keys:
        if key != '':
            cluster_sizes[key] = 0
    return cluster_sizes  

def save_cluster_size_data(num_nodes, graph_type, appeal, filename, cluster_sizes):
    print("Svaing Data")
    wb = openpyxl.load_workbook(FILE_DIRECTORY_PREFIX + 'Cluster_Data_No_Neg.xlsx')
    sheets = wb.sheetnames
    cluster_size_data = wb[sheets[0]]
    counts = [value for key, value in cluster_sizes.items()]
    cluster_size_data.append([num_nodes] + [appeal] + [graph_type] + [filename] + counts)
    wb.save(FILE_DIRECTORY_PREFIX + 'Cluster_Data_No_Neg.xlsx')

def write_header(lines):
    print("Write Header")
    wb = openpyxl.load_workbook(FILE_DIRECTORY_PREFIX + 'Cluster_Data_No_Neg.xlsx')
    sheets = wb.sheetnames
    cluster_size_data = wb[sheets[0]]
    cluster_size_data.append(["Num Nodes","Appeal","Graph Type","Location"] + lines)
    wb.save(FILE_DIRECTORY_PREFIX + 'Cluster_Data_No_Neg.xlsx')

def histogram_data_collection(node_sizes, runs, appeals):
    get_all_possible_cluster_sizes(node_sizes, runs, appeals)
    record_occurences_of_cluster_sizes(node_sizes, runs, appeals)

def test_facebook(runs, k_vals, appeals):
    for appeal in appeals:
        for k in k_vals:
            for i in range(runs):
                filename = "facebook_combined.txt"
                driver.test_file(filename, k, appeal)

def main():
    node_sizes = [5000] # ["500", "1000", "2000", "5000"] | ["150", "500", "1000", "2000"]
    node_offset = [3]
    k_vals = ["100"] #["30","40","60","70","80","90"] #"10","20","50","100"] # ["5","10","20","50"]
    appeals = ["0.25", "0.5", "0.75"] # ["0.5"]#["0.25", "0.5", "0.75"] # ["0.5", "0.5", "0.5", "0.5"]
    runs = 25

    m = 2
    p = 0.2

    # if not is_ID_reset():
    #     do_gen = input("Do you want to generate new files? Enter [Y] or [N]: ")
    #     if do_gen == "Y":
    #         generate_and_save_graphs(node_sizes, runs, m=2,p=0.2)
    record_cluster_data(node_sizes, node_offset, runs, k_vals, appeals)
    #test_facebook(runs, k_vals, appeals)
    # test_dif_combos(node_sizes, node_offset, runs, k_vals, appeals)

    # histogram_data_collection([5000], runs, appeals)
record_occurences_of_cluster_sizes()
# main()











# get_cluster_data()
# driver.test_file("1", "5", "0.5")

# # the functions below were lazy ways to generate and test programs, relying on generation while testing
# def generate_all_possible_combos():
#     node_sizes = ["500", "1000", "2000", "5000"] # ["150", "500", "1000", "2000"]
#     possible_k = ["10","20","50","100"] # ["5","10","20","50"]
#     possible_criticalities = ["0.5", "0.75"] # ["0.5", "0.5", "0.5", "0.5"]

#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     runs = 25

#     for criticality in possible_criticalities:
#         for num_nodes in node_sizes:
#             for k in possible_k:
#                 for i in range(runs):
#                     print("RUN " + str(i) + ": " + str(num_nodes) + " " + str(k) + " " + str(criticality))
#                     driver.test_new_file(num_nodes, k, criticality, do_remove_cycles, do_assumption_1)

# def test():
#     graph_types = ["BA"]
#     offset = [1]
#     node_size = ["150"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             size_offset = 0
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# def practice():
#     num_nodes = "500"
#     k = "10"
#     criticality = "0.5"
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for i in range(6):
#         print("RUN " + str(i) + ": " + str(num_nodes) + " " + str(k) + " " + str(criticality))
#         driver.test_new_file(num_nodes, k, criticality, do_remove_cycles, do_assumption_1)
# def retest_all_ER_WS():
#     graph_types = ["ER","WS"]
#     offset = [2,3]
#     node_size = ["150", "500", "1000", "2000"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# def retest_all_files():
#     graph_types = ["BA","ER","WS"]
#     offset = [1,2,3]
#     node_size = ["150", "500", "1000", "2000"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# def retest_BA():
#     graph_types = ["BA"]
#     offset = [1]
#     node_size = ["150", "500", "1000", "2000"]
#     do_remove_cycles = "False"
#     do_assumption_1 = "False"

#     for graph, ID_off in zip(graph_types, offset):
#         for size_offset, size in enumerate(node_size):
#             for i in range(50):
#                 filename = graph + "/" + size + "/" + str((size_offset * 150) + (i * 3) + ID_off)
#                 # if filename[:3] != "BA/":
#                 driver.retest_old_file(filename, do_remove_cycles, do_assumption_1)

# # practice()